# -*- coding: utf-8 -*-
"""
Importador del CSV "Merian delivery transactions" exportado desde Veridapt.

Lee el CSV, normaliza cada fila al formato de la hoja
"delivery_transaction_213_202312" del Excel historico y la inserta como
una fila nueva al final de la tabla "Table1".  Se evitan duplicados
comparando (Collected At, Tank, Volume) contra las filas ya presentes.

Las dos ultimas columnas (Variance y %) se escriben como formulas que
hacen referencia a la propia tabla, igual que las filas ya existentes:
    =Table1[[#This Row],[Docket Volume]]-Table1[[#This Row],[Volume]]
    =Table1[[#This Row],[Variance]]/Table1[[#This Row],[Docket Volume]]

Uso desde la app:
    parsed = parse_delivery_csv(path)
    summary = append_to_excel(excel_path, parsed)
"""
from __future__ import annotations

import csv
import datetime
import os
import re

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

SHEET = "delivery_transaction_213_202312"
TABLE = "Table1"

# Columnas de Table1 (indices 1-based).
COL_PRODUCT = 1
COL_DATE = 2
COL_TANK = 3
COL_DOCKET = 4
COL_SUPPLIER = 5
COL_CONFIRMED = 6
COL_TYPE = 7
COL_VOLUME = 8
COL_DOCKET_VOLUME = 9
COL_VARIANCE = 10
COL_PCT = 11

# Mapeo Tank/Product CSV -> Product del Excel.  Se busca por substring del
# nombre del tanque porque es lo mas estable; el campo Product del CSV
# usa otros nombres ('SAE 30', 'Hydraulic Fluid', etc.) que no coinciden.
_PRODUCT_RULES = [
    ("S4CX10W", "Spirax S4CX10W"),
    ("S4CX30",  "Spirax S4CX30"),
    ("S5CFDM60","Spirax S5CFDM60"),
    ("S3M46",   "Tellus S3M46"),
    ("Rimula",  "15W40"),
    ("15W40",   "15W40"),
]


def _to_float(s):
    """Convierte '3,090.36' -> 3090.36.  Devuelve None si esta vacio."""
    if s is None:
        return None
    s = str(s).strip()
    if not s:
        return None
    # Las celdas del CSV vienen entrecomilladas; csv.reader ya quita las
    # comillas, pero por si acaso:
    s = s.strip('"').replace(",", "")
    try:
        return float(s)
    except ValueError:
        return None


def _parse_date(s):
    """'28/04/2026 10:17' -> datetime.  None si no se puede parsear."""
    if not s:
        return None
    s = s.strip().strip('"')
    for fmt in ("%d/%m/%Y %H:%M", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y"):
        try:
            return datetime.datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _product_from_tank(tank: str, csv_product: str) -> str | None:
    """Devuelve el nombre de producto del Excel a partir del nombre del
    tanque (campo 'Delivery To' del CSV).  Se intenta primero con el
    tanque, luego con el campo Product como fallback."""
    if tank:
        for needle, name in _PRODUCT_RULES:
            if needle.lower() in tank.lower():
                return name
    if csv_product:
        for needle, name in _PRODUCT_RULES:
            if needle.lower() in csv_product.lower():
                return name
    return None


def _confirmed(status: str) -> str:
    s = (status or "").strip().lower()
    if s == "all_ok":
        return "Yes"
    return "No"


def _type(t: str) -> str:
    t = (t or "").strip()
    if not t:
        return ""
    if t.upper() == "GAUGED":
        return "Gauged"
    return t.capitalize()


def parse_delivery_csv(path: str) -> list[dict]:
    """Devuelve una lista de dicts (uno por fila valida del CSV), con las
    claves que se van a escribir al Excel: product, date, tank, docket,
    confirmed, type, volume, docket_volume."""
    rows = []
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for raw in reader:
            tank = (raw.get("Delivery To") or "").strip()
            date = _parse_date(raw.get("Date"))
            volume = _to_float(raw.get("Volume"))
            if not tank or date is None or volume is None:
                # Sin tanque / fecha / volumen no se puede registrar.
                continue
            product = _product_from_tank(tank, raw.get("Product") or "")
            if product is None:
                # Producto no reconocido (p.ej. otros lubricantes que no
                # entran en el reporte semanal).
                continue
            rows.append({
                "product":       product,
                "date":          date,
                "tank":          tank,
                "docket":        (raw.get("Delivery Docket Number")
                                  or "").strip() or None,
                "confirmed":     _confirmed(raw.get("Confirmation Status")),
                "type":          _type(raw.get("Transaction Type")),
                "volume":        volume,
                "docket_volume": _to_float(raw.get("Field Entered Volume")),
            })
    return rows


def _existing_keys(ws, header_row: int = 1) -> set[tuple]:
    """Devuelve el set de (date, tank, volume) ya presentes en la hoja,
    para detectar duplicados."""
    keys = set()
    for ri in range(header_row + 1, ws.max_row + 1):
        d = ws.cell(row=ri, column=COL_DATE).value
        t = ws.cell(row=ri, column=COL_TANK).value
        v = ws.cell(row=ri, column=COL_VOLUME).value
        if d is None and t is None and v is None:
            continue
        keys.add((d, t, round(float(v), 2) if isinstance(v, (int, float))
                 else v))
    return keys


def _expand_table_ref(ws, new_last_row: int) -> None:
    """Actualiza el atributo 'ref' de Table1 para que incluya las filas
    nuevas (sino Excel ignora las filas al abrir el archivo)."""
    table = ws.tables.get(TABLE)
    if table is None:
        return
    # ref viene como 'A1:K360'.  Solo cambiamos la parte de fila final.
    m = re.match(r"^([A-Z]+)(\d+):([A-Z]+)(\d+)$", table.ref)
    if not m:
        return
    c1, r1, c2, _ = m.groups()
    table.ref = "%s%s:%s%d" % (c1, r1, c2, new_last_row)


def append_to_excel(excel_path: str, parsed_rows: list[dict]) -> dict:
    """Escribe `parsed_rows` al final de Table1 en `excel_path`, saltando
    duplicados.  Devuelve un dict con resumen: inserted, skipped, total."""
    wb = load_workbook(excel_path)
    if SHEET not in wb.sheetnames:
        raise ValueError("El Excel no contiene la hoja '%s'." % SHEET)
    ws = wb[SHEET]
    existing = _existing_keys(ws)

    inserted = 0
    skipped = 0
    next_row = ws.max_row + 1
    for r in parsed_rows:
        key = (r["date"], r["tank"], round(r["volume"], 2))
        if key in existing:
            skipped += 1
            continue
        ws.cell(row=next_row, column=COL_PRODUCT,       value=r["product"])
        ws.cell(row=next_row, column=COL_DATE,          value=r["date"])
        ws.cell(row=next_row, column=COL_TANK,          value=r["tank"])
        ws.cell(row=next_row, column=COL_DOCKET,        value=r["docket"])
        ws.cell(row=next_row, column=COL_SUPPLIER,      value=None)
        ws.cell(row=next_row, column=COL_CONFIRMED,     value=r["confirmed"])
        ws.cell(row=next_row, column=COL_TYPE,          value=r["type"])
        ws.cell(row=next_row, column=COL_VOLUME,        value=r["volume"])
        ws.cell(row=next_row, column=COL_DOCKET_VOLUME, value=r["docket_volume"])
        ws.cell(row=next_row, column=COL_VARIANCE,
                value="=Table1[[#This Row],[Docket Volume]]"
                      "-Table1[[#This Row],[Volume]]")
        ws.cell(row=next_row, column=COL_PCT,
                value="=Table1[[#This Row],[Variance]]"
                      "/Table1[[#This Row],[Docket Volume]]")
        # Formatos numericos identicos a los de las filas existentes para
        # que Volume / Variance se muestren con 2 decimales y la columna %
        # como porcentaje (sino el 0.10057 se ve crudo en vez de '10,06%').
        ws.cell(row=next_row, column=COL_DATE).number_format = \
            "dd/mm/yyyy hh:mm"
        ws.cell(row=next_row, column=COL_VOLUME).number_format = "0.00"
        ws.cell(row=next_row, column=COL_VARIANCE).number_format = "#,##0.00"
        ws.cell(row=next_row, column=COL_PCT).number_format = "0.00%"
        existing.add(key)
        inserted += 1
        next_row += 1

    # Normalizar formatos en TODAS las filas de datos.  Esto corrige
    # tambien las filas insertadas por versiones anteriores que quedaron
    # en formato 'General' (sin '%') y garantiza consistencia visual.
    for ri in range(2, next_row):
        ws.cell(row=ri, column=COL_VOLUME).number_format = "0.00"
        ws.cell(row=ri, column=COL_DOCKET_VOLUME).number_format = "General"
        ws.cell(row=ri, column=COL_VARIANCE).number_format = "#,##0.00"
        ws.cell(row=ri, column=COL_PCT).number_format = "0.00%"

    if inserted:
        _expand_table_ref(ws, next_row - 1)
    wb.save(excel_path)
    return {"inserted": inserted, "skipped": skipped,
            "total": len(parsed_rows)}
