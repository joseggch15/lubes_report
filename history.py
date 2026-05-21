# -*- coding: utf-8 -*-
"""
Lectura del Excel historico "Reconciliation Lubs Weekly Ver2.xlsx".

Ese Excel contiene una hoja "Recon <producto>" por cada producto, con UNA
fila por semana (columna Date) y, para algunos productos, varias filas por
fecha (un tanque cada una).

Este modulo indexa esa informacion por fecha para que la interfaz pueda:
  - ofrecer la lista de fechas disponibles,
  - extraer los datos de una semana concreta,
  - avisar cuando una fecha no tiene informacion registrada.
"""
from __future__ import annotations

import datetime

import openpyxl

import report_model as m

# Productos que tienen seccion de "Tank log" / "Reconciliation" en el
# reporte (Figuras 9 a 18). Incluye S4CX10W, que no esta en la tabla
# consolidada pero si tiene sus dos figuras.
TANK_LOG_PRODUCTS = ["S4CX30", "S4CX10W", "15W40", "S5CFDM60", "Tellus S3M46"]

# Columna "Recon %" en las hojas "WeeklyVariance ...".
WV_COL_DATE = 0
WV_COL_RECON = 2

# Algunos productos tienen varios tanques en su hoja "Recon ...". Para esos
# casos se indica de cual tanque (columna Site) se debe tomar el dato. Si un
# producto no figura aqui, se suman todas sus filas de la fecha.
PRODUCT_TANK_FILTER = {
    "S4CX30": "Tank 21",
}

# Columnas (0-indexadas) de las hojas "Recon <producto>".
COL_DATE = 0
COL_SITE = 1
COL_OPENING = 2
COL_DELIVERIES = 3
COL_TRANSACTIONS = 4
COL_CLOSING = 7
COL_TO_EQUIPMENT = 11
COL_OTHER = 12
COL_TRANSFERS = 13


# Columnas de la hoja "delivery_transaction_...".
DT_PRODUCT = 0
DT_COLLECTED = 1
DT_CONFIRMED = 5
DT_VOLUME = 7
DT_DOCKET_VOLUME = 8


def _match_delivery_product(text) -> str | None:
    """Mapea el texto de producto de la hoja de transacciones al key del
    reporte. Devuelve None para productos ajenos al reporte.

    Caso especial: los tickets con producto 'Spirax S4CX10W' se reportan
    como entregas de 'S4CX30' porque ese tanque almacena el mismo producto
    (S4CX30) — el W solo indica el tanque, no un lubricante distinto. Esta
    regla aplica EXCLUSIVAMENTE a la tabla de tickets de entrega; en las
    hojas Recon, en los Tank Logs y en las tendencias WeeklyVariance,
    S4CX10W sigue siendo su propio producto.
    """
    if not text:
        return None
    up = str(text).upper().replace(" ", "")
    if "S4CX10W" in up:
        return "S4CX30"
    if "S4CX30" in up:
        return "S4CX30"
    if "15W40" in up or "RIMULA" in up:
        return "15W40"
    if "S5CFDM60" in up:
        return "S5CFDM60"
    if "TELLUS" in up or "S3M46" in up:
        return "Tellus S3M46"
    return None


def _match_wv_sheet(product_key: str, sheet_names: list) -> str | None:
    """Asocia un producto con su hoja 'WeeklyVariance ...' del Excel."""
    key = product_key.lower().replace(" ", "")
    for name in sheet_names:
        compact = name.lower().replace(" ", "")
        if not compact.startswith("weeklyvariance"):
            continue
        if key == "s4cx30" and "s4cx30" in compact and "s4cx10w" not in compact:
            return name
        if key == "s4cx10w" and "s4cx10w" in compact:
            return name
        if key == "15w40" and "15w40" in compact:
            return name
        if key == "s5cfdm60" and "s5cfdm60" in compact:
            return name
        if key.startswith("tellus") and "tellus" in compact:
            return name
    return None


def _match_sheet(product_key: str, sheet_names: list) -> str | None:
    """Asocia un producto del reporte con su hoja 'Recon ...' del Excel."""
    recon = [s for s in sheet_names if s.lower().startswith("recon")]
    key = product_key.lower().replace(" ", "")
    for name in recon:
        compact = name.lower().replace(" ", "")
        if key == "s4cx30" and "s4cx30" in compact and "s4cx10w" not in compact:
            return name
        if key == "15w40" and "15w40" in compact:
            return name
        if key == "s5cfdm60" and "s5cfdm60" in compact:
            return name
        if key.startswith("tellus") and "tellus" in compact:
            return name
    return None


class WeekRecord:
    """Datos agregados de un producto para una semana concreta."""

    __slots__ = ("found", "opening", "deliveries", "transactions",
                 "closing", "to_equipment", "other", "transfers", "tanks")

    def __init__(self):
        self.found = False
        self.opening = 0.0
        self.deliveries = 0.0
        self.transactions = 0.0
        self.closing = 0.0
        self.to_equipment = 0.0
        self.other = 0.0
        self.transfers = 0.0
        self.tanks = []


class HistoryStore:
    """Indexa el Excel historico por producto y por fecha."""

    def __init__(self):
        # {product_key: {date(datetime.date): WeekRecord}}
        self._by_product: dict = {}
        # lista de tickets: {product, dt(datetime), volume, docket}
        self._deliveries: list = []
        # {product_key: [(date, recon_pct), ...]}  (hojas WeeklyVariance)
        self._weekly_variance: dict = {}
        self.source_path = None

    # -- carga --------------------------------------------------------------

    def load(self, path: str) -> None:
        wb = openpyxl.load_workbook(path, data_only=True)
        names = wb.sheetnames
        self._by_product = {}
        self.source_path = path

        for product_key in m.PRODUCT_KEYS:
            sheet_name = _match_sheet(product_key, names)
            week_map: dict = {}
            tank_filter = PRODUCT_TANK_FILTER.get(product_key)
            if sheet_name is not None:
                ws = wb[sheet_name]
                for row in ws.iter_rows(min_row=3, values_only=True):
                    raw_date = row[COL_DATE] if row else None
                    if not isinstance(raw_date, datetime.datetime):
                        continue
                    day = raw_date.date()
                    rec = week_map.setdefault(day, WeekRecord())
                    # Una fila "vacia" (solo fecha) no aporta datos.
                    if row[COL_OPENING] is None and row[COL_CLOSING] is None:
                        continue
                    # Productos con varios tanques: solo se toma el indicado.
                    if tank_filter is not None:
                        site = row[COL_SITE]
                        if (site is None or str(site).strip().lower()
                                != tank_filter.strip().lower()):
                            continue
                    rec.found = True
                    rec.opening += m._num(row[COL_OPENING])
                    rec.deliveries += m._num(row[COL_DELIVERIES])
                    rec.transactions += m._num(row[COL_TRANSACTIONS])
                    rec.closing += m._num(row[COL_CLOSING])
                    rec.to_equipment += m._num(row[COL_TO_EQUIPMENT])
                    rec.other += m._num(row[COL_OTHER])
                    rec.transfers += m._num(row[COL_TRANSFERS])
                    if row[COL_SITE]:
                        rec.tanks.append(str(row[COL_SITE]))
            self._by_product[product_key] = week_map

        # Tickets de entrega (hoja "delivery_transaction_...").
        self._deliveries = []
        dt_sheet = next((s for s in names
                         if s.lower().startswith("delivery_transaction")), None)
        if dt_sheet is not None:
            ws = wb[dt_sheet]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue
                key = _match_delivery_product(row[DT_PRODUCT])
                collected = row[DT_COLLECTED]
                if key is None or not isinstance(collected, datetime.datetime):
                    continue
                # Solo entregas confirmadas (las del reporte).
                if str(row[DT_CONFIRMED]).strip().lower() != "yes":
                    continue
                self._deliveries.append({
                    "product": key,
                    "dt": collected,
                    "volume": m._num(row[DT_VOLUME]),
                    "docket": m._num(row[DT_DOCKET_VOLUME]),
                })

        # Tendencia semanal de variance (hojas "WeeklyVariance ...").
        self._weekly_variance = {}
        for product_key in TANK_LOG_PRODUCTS:
            wv_name = _match_wv_sheet(product_key, names)
            points = []
            if wv_name is not None:
                for row in wb[wv_name].iter_rows(min_row=3, values_only=True):
                    if not row:
                        continue
                    raw_date = row[WV_COL_DATE]
                    recon = row[WV_COL_RECON] if len(row) > WV_COL_RECON else None
                    if isinstance(raw_date, datetime.datetime) and recon is not None:
                        points.append((raw_date.date(), m._num(recon)))
            self._weekly_variance[product_key] = points

    # -- consultas ----------------------------------------------------------

    def is_loaded(self) -> bool:
        return bool(self._by_product)

    def available_dates(self) -> list:
        """Fechas (datetime.date) que tienen datos en al menos un producto,
        ordenadas de la mas reciente a la mas antigua."""
        dates = set()
        for week_map in self._by_product.values():
            for day, rec in week_map.items():
                if rec.found:
                    dates.add(day)
        return sorted(dates, reverse=True)

    def lookup(self, product_key: str, day: datetime.date) -> WeekRecord | None:
        """Devuelve el WeekRecord de un producto/fecha, o None si no existe
        informacion registrada para esa combinacion."""
        rec = self._by_product.get(product_key, {}).get(day)
        if rec is None or not rec.found:
            return None
        return rec

    def week(self, day: datetime.date) -> dict:
        """Devuelve {product_key: WeekRecord | None} para toda una semana."""
        return {key: self.lookup(key, day) for key in m.PRODUCT_KEYS}

    def _week_window(self, day: datetime.date) -> tuple:
        """Devuelve (inicio_exclusivo, fin_inclusivo) de la semana que termina
        en `day`. El inicio es la fecha de semana anterior disponible; si no
        hay, se usan 7 dias hacia atras."""
        prev = None
        for d in self.available_dates():   # ordenadas de reciente a antigua
            if d < day and (prev is None or d > prev):
                prev = d
        if prev is None:
            prev = day - datetime.timedelta(days=7)
        return prev, day

    def recon_trend(self, product_key: str, day: datetime.date,
                    limit: int = 12) -> list:
        """Devuelve los ultimos `limit` puntos (date, recon_pct) de la hoja
        WeeklyVariance del producto, hasta la fecha `day` inclusive."""
        points = [p for p in self._weekly_variance.get(product_key, [])
                  if p[0] <= day]
        return points[-limit:]

    def period_for(self, day: datetime.date) -> tuple:
        """Devuelve (period_full, cover_date) derivados de la semana que
        termina en `day`. Asi el texto del periodo y la fecha de portada
        siempre coinciden con la fecha seleccionada."""
        start, end = self._week_window(day)
        return (m.format_period(start, end), m.format_cover_date(end))

    def deliveries_for_week(self, day: datetime.date) -> dict:
        """Devuelve {product_key: [ {date, volume, docket}, ... ]} con los
        tickets de entrega confirmados de la semana que termina en `day`."""
        start, end = self._week_window(day)
        out = {key: [] for key in m.PRODUCT_KEYS}
        for tk in sorted(self._deliveries, key=lambda x: x["dt"]):
            d = tk["dt"].date()
            if start < d <= end:
                out[tk["product"]].append({
                    "date": tk["dt"].strftime("%d/%m/%Y"),
                    "volume": tk["volume"],
                    "docket": tk["docket"],
                })
        return out


def apply_week_to_data(data: dict, store: HistoryStore,
                       day: datetime.date) -> tuple:
    """Vuelca los datos historicos de la fecha `day` sobre la estructura
    `data` (la del report_model). Devuelve (encontrados, faltantes):
    listas de claves de producto con y sin informacion registrada.

    Se actualizan consolidada, dispensing y la tabla de tickets de entrega.
    Los textos editables se conservan.
    """
    found, missing = [], []
    week = store.week(day)
    by_site = {r["site"]: r for r in data["consolidated"]}
    deliveries = store.deliveries_for_week(day)

    # Tendencia de variance para los graficos "Reconciliation for the
    # product" (Figuras 10, 12, 14, 16, 18).
    data["recon_trends"] = {
        key: store.recon_trend(key, day) for key in TANK_LOG_PRODUCTS}

    for product_key in m.PRODUCT_KEYS:
        rec = week.get(product_key)
        if rec is None:
            missing.append(product_key)
        else:
            found.append(product_key)
            row = by_site.get(product_key)
            if row is not None:
                row["opening"] = round(rec.opening, 2)
                row["deliveries"] = round(rec.deliveries, 2)
                row["transactions"] = round(rec.transactions, 2)
                row["closing"] = round(rec.closing, 2)
            prod = data["products"].setdefault(product_key, {})
            prod["disp_equipment"] = round(rec.to_equipment, 2)
            prod["disp_other"] = round(rec.other, 2)
            prod["disp_transfers"] = round(rec.transfers, 2)
        # Tickets de entrega de la semana (puede ser lista vacia).
        data["deliveries"][product_key] = deliveries.get(product_key, [])

    return found, missing
