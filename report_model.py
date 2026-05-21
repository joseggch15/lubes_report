# -*- coding: utf-8 -*-
"""
Modelo de datos y logica de generacion del
"Weekly Lubes tank reconciliation report".

Responsabilidades:
  - Definir la plantilla de Excel de entrada.
  - Leer / escribir ese Excel.
  - Calcular las columnas derivadas (variance, %, totales).
  - Construir el contexto para docxtpl y renderizar el .docx final.

No contiene interfaz grafica (ver app.py).
"""
from __future__ import annotations

import os
import shutil
import tempfile

import openpyxl
from docxtpl import DocxTemplate, InlineImage, RichText
from docx.shared import Mm

import charts

# --------------------------------------------------------------------------
# Configuracion fija del reporte
# --------------------------------------------------------------------------

# Orden oficial de los productos (igual al de la tabla consolidada del Word).
PRODUCT_KEYS = ["S4CX30", "15W40", "S5CFDM60", "Tellus S3M46"]

# Nombre "bonito" que aparece en los textos del reporte.
PRODUCT_NAMES = {
    "S4CX30": "SPIRAX S4CX30",
    "15W40": "Rimula R4X15W40",
    "S5CFDM60": "Spirax S5CFDM60",
    "Tellus S3M46": "Hydraulic Fluid - Tellus S3M46",
}

# Consideraciones por defecto (texto base reutilizado cada semana).
DEFAULT_CONSIDERATIONS = {
    "S4CX30": [
        "Delivery volume has been corrected to match the ticket volume.",
        "It is recommended to carry out research into deliveries with this "
        "product, especially because the consumption projection does not "
        "correspond to what is felt, this can be observed from the Storage graphs.",
    ],
    "15W40": [
        "Delivery volume has been corrected to match the ticket volume.",
        "The value is assigned as 0% because there is no inlet flowmeter.",
        "In service trucks lube tanks distribution may have changed, we cannot "
        "know with certainty the sensed volume that corresponds to a certain lube.",
        "Overfilling the tank above the safe fill level is causing erroneous "
        "data to be considered for reconciliation.",
    ],
    "S5CFDM60": [
        "Delivery volume has been corrected to match the ticket volume.",
        "The value is assigned as 0% because there is no inlet flowmeter.",
        "In service trucks lube tanks distribution may have changed, we cannot "
        "know with certainty the sensed volume that corresponds to a certain lube.",
        "Overfilling the tank above the safe fill level is causing erroneous "
        "data to be considered for reconciliation.",
    ],
    "Tellus S3M46": [
        "Delivery volume has been corrected to match the ticket volume.",
        "The value is assigned as 0% because there is no inlet flowmeter.",
        "In service trucks lube tanks distribution may have changed, we cannot "
        "know with certainty the sensed volume that corresponds to a certain lube.",
        "Overfilling the tank above the safe fill level is causing erroneous "
        "data to be considered for reconciliation.",
    ],
}

# Figuras: 18 figuras del cuerpo + 1 imagen de la tabla de tareas.
FIGURE_KEYS = ["fig%d" % i for i in range(1, 19)] + ["fig_tasks"]

# Productos con seccion "Tank log" / "Reconciliation" (Figuras 9-18), en el
# orden en que aparecen en el reporte. Incluye S4CX10W.
TANK_LOG_ORDER = ["S4CX30", "S4CX10W", "15W40", "S5CFDM60", "Tellus S3M46"]

# Nombre para mostrar en los titulos de los graficos de tanque.
TANK_DISPLAY = {
    "S4CX30": "Spirax S4CX30",
    "S4CX10W": "Spirax S4CX10W",
    "15W40": "Rimula R4X15W40",
    "S5CFDM60": "Spirax S5CFDM60",
    "Tellus S3M46": "Tellus S3M46",
}

# Titulo del grafico Tank Log que aparece en el original (se exporta de Excel).
TANK_TITLES = {
    "S4CX30": "Tank WS - Spirax S4CX30 - Tank 2",
    "S4CX10W": "Tank WS - Spirax S4CX10W - Tank 1",
    "15W40": "Tank WS - Rimula R4X15W40 - Tank 6",
    "S5CFDM60": "Tank WS - Spirax S5CFDM60 - Tank 4",
    "Tellus S3M46": "Tank WS - Hydraulic Fluid - Tellus S3M46 - Tank 9",
}

# Fuente que usa el reporte original.
REPORT_FONT = "Arial"

# Etiqueta descriptiva de cada figura (se muestra en la interfaz grafica).
FIGURE_LABELS = {
    "fig1": "Figura 1  - Recon. con tickets SOL - S4CX30",
    "fig2": "Figura 2  - Recon. con tickets SOL - 15W40",
    "fig3": "Figura 3  - Recon. con tickets SOL - S5CFDM60",
    "fig4": "Figura 4  - Recon. con tickets SOL - Tellus S3M46",
    "fig5": "Figura 5  - Distribucion de transacciones - S4CX30",
    "fig6": "Figura 6  - Distribucion de transacciones - 15W40",
    "fig7": "Figura 7  - Distribucion de transacciones - S5CFDM60",
    "fig8": "Figura 8  - Distribucion de transacciones - Tellus S3M46",
    "fig9": "Figura 9  - Tank Log - Spirax S4CX30",
    "fig10": "Figura 10 - Reconciliation - S4CX30",
    "fig11": "Figura 11 - Tank Log - Spirax S4CX10W",
    "fig12": "Figura 12 - Reconciliation - S4CX10W",
    "fig13": "Figura 13 - Tank Log - Rimula 15W40",
    "fig14": "Figura 14 - Reconciliation - Rimula 15W40",
    "fig15": "Figura 15 - Tank Log - S5CFDM60",
    "fig16": "Figura 16 - Reconciliation - S5CFDM60",
    "fig17": "Figura 17 - Tank Log - Tellus S3M46",
    "fig18": "Figura 18 - Reconciliation - Tellus S3M46",
    "fig_tasks": "Tabla 6  - Resumen de tareas de la semana (imagen)",
}

# Ancho con el que se inserta cada figura en el Word. Se eligio para que
# coincida con el tamano que ocupaban las imagenes pegadas desde Excel en
# los reportes hechos a mano (medido en M0518: ~104mm las barras de
# entregas, ~95mm las donas, ~158mm los tank logs, ~150mm los recon trends).
_DELIVERY_BAR_FIGS = {"fig1", "fig2", "fig3", "fig4"}
_PIE_FIGS = {"fig5", "fig6", "fig7", "fig8"}
_RECON_TREND_FIGS = {"fig10", "fig12", "fig14", "fig16", "fig18"}


def figure_width(fig_key: str) -> Mm:
    if fig_key in _DELIVERY_BAR_FIGS:
        return Mm(105)
    if fig_key in _PIE_FIGS:
        return Mm(95)
    if fig_key in _RECON_TREND_FIGS:
        return Mm(150)
    # Tank logs (fig9, 11, 13, 15, 17) y fig_tasks.
    return Mm(158)


# --------------------------------------------------------------------------
# Utilidades de numeros y formato
# --------------------------------------------------------------------------

def _num(value, default=0.0) -> float:
    """Convierte cualquier celda de Excel a float de forma tolerante."""
    if value is None or value == "":
        return float(default)
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("%", "")
    # Acepta formato europeo (1.234,56) y anglosajon (1,234.56).
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return float(default)


def fmt_int(value) -> str:
    """Entero con punto como separador de miles: 16576 -> '16.576'."""
    n = round(_num(value))
    return "{:,}".format(int(n)).replace(",", ".")


def fmt_pct(value, decimals=1) -> str:
    """Porcentaje con coma decimal: -4.11 -> '-4,11%'."""
    text = ("{:." + str(decimals) + "f}").format(_num(value))
    return text.replace(".", ",") + "%"


_MONTHS_EN = ["January", "February", "March", "April", "May", "June", "July",
              "August", "September", "October", "November", "December"]


def ordinal(n: int) -> str:
    """Numero con sufijo ordinal en ingles: 1 -> '1st', 18 -> '18th'."""
    n = int(n)
    if 10 <= n % 100 <= 20:
        suffix = "th"
    else:
        suffix = {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
    return "%d%s" % (n, suffix)


def format_cover_date(d) -> str:
    """Fecha de portada: date(2026,5,18) -> 'May 18th, 2026'."""
    return "%s %s, %d" % (_MONTHS_EN[d.month - 1], ordinal(d.day), d.year)


def format_period(start, end) -> str:
    """Texto de periodo: -> 'April 30th to May 6th, 2026'."""
    return "%s %s to %s %s, %d" % (
        _MONTHS_EN[start.month - 1], ordinal(start.day),
        _MONTHS_EN[end.month - 1], ordinal(end.day), end.year)


# --------------------------------------------------------------------------
# Estructura de datos por defecto
# --------------------------------------------------------------------------

def default_data() -> dict:
    """Devuelve la estructura de datos completa con valores de ejemplo
    (los del reporte del 11 al 18 de mayo de 2026)."""
    consolidated_seed = {
        "S4CX30": (16576, 11417, 8718, 18918),
        "15W40": (23843, 0, 3317, 20571),
        "S5CFDM60": (4218, 3683, 1630, 6105),
        "Tellus S3M46": (3980, 4225, 1658, 6372),
    }
    dispensing_seed = {
        "S4CX30": (6732, 0, 1986),
        "15W40": (1823, 0, 1494),
        "S5CFDM60": (1630, 0, 0),
        "Tellus S3M46": (0, 0, 1658),
    }
    deliveries_seed = {
        "S4CX30": [("15/05/2026", 6040, 5700), ("15/05/2026", 5377, 5300)],
        "15W40": [("11/05/2026 - 18/05/2026", 0, 0)],
        "S5CFDM60": [("14/05/2026", 3683, 3500)],
        "Tellus S3M46": [("13/05/2026", 4225, 4000)],
    }

    consolidated = []
    for key in PRODUCT_KEYS:
        opening, deliveries, transactions, closing = consolidated_seed[key]
        consolidated.append({
            "site": key, "opening": opening, "deliveries": deliveries,
            "transactions": transactions, "closing": closing,
        })

    products = {}
    for key in PRODUCT_KEYS:
        eq, ot, tr = dispensing_seed[key]
        products[key] = {
            "recon_sentence": "",          # se autogenera si queda vacio
            "considerations": list(DEFAULT_CONSIDERATIONS[key]),
            "delivery_narrative": "",      # se autogenera si queda vacio
            "issued": "",                  # se autogenera si queda vacio
            "disp_equipment": eq,
            "disp_other": ot,
            "disp_transfers": tr,
        }

    deliveries = {}
    for key in PRODUCT_KEYS:
        deliveries[key] = [
            {"date": d, "volume": v, "docket": k}
            for (d, v, k) in deliveries_seed[key]
        ]

    return {
        "period_full": "May 11th to 18th, 2026",
        "cover_date": "May 18th, 2026",
        "consolidated": consolidated,
        "products": products,
        "deliveries": deliveries,
    }


# --------------------------------------------------------------------------
# Lectura / escritura del Excel de entrada
# --------------------------------------------------------------------------

SHEET_META = "Meta"
SHEET_CONSOLIDATED = "Consolidated"
SHEET_PRODUCTS = "Products"
SHEET_DELIVERIES = "Deliveries"


def save_excel(data: dict, path: str) -> None:
    """Escribe la estructura de datos a un archivo Excel de entrada."""
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = SHEET_META
    ws.append(["Campo", "Valor"])
    ws.append(["period_full", data["period_full"]])
    ws.append(["cover_date", data.get("cover_date", "")])

    ws = wb.create_sheet(SHEET_CONSOLIDATED)
    ws.append(["site", "opening", "deliveries", "transactions", "closing"])
    for row in data["consolidated"]:
        ws.append([row["site"], row["opening"], row["deliveries"],
                   row["transactions"], row["closing"]])

    ws = wb.create_sheet(SHEET_PRODUCTS)
    ws.append(["key", "recon_sentence", "considerations", "delivery_narrative",
               "issued", "disp_equipment", "disp_other", "disp_transfers"])
    for key in PRODUCT_KEYS:
        p = data["products"][key]
        ws.append([key, p["recon_sentence"],
                   "\n".join(p["considerations"]),
                   p["delivery_narrative"], p["issued"],
                   p["disp_equipment"], p["disp_other"], p["disp_transfers"]])

    ws = wb.create_sheet(SHEET_DELIVERIES)
    ws.append(["product_key", "date", "volume", "docket"])
    for key in PRODUCT_KEYS:
        for d in data["deliveries"][key]:
            ws.append([key, d["date"], d["volume"], d["docket"]])

    wb.save(path)


def load_excel(path: str) -> dict:
    """Lee un Excel de entrada y devuelve la estructura de datos."""
    wb = openpyxl.load_workbook(path, data_only=True)
    data = default_data()

    if SHEET_META in wb.sheetnames:
        for row in wb[SHEET_META].iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            if row[0] == "period_full" and row[1]:
                data["period_full"] = str(row[1])
            elif row[0] == "cover_date" and row[1]:
                data["cover_date"] = str(row[1])

    if SHEET_CONSOLIDATED in wb.sheetnames:
        consolidated = []
        for row in wb[SHEET_CONSOLIDATED].iter_rows(min_row=2, values_only=True):
            if not row or row[0] in (None, ""):
                continue
            consolidated.append({
                "site": str(row[0]),
                "opening": _num(row[1]), "deliveries": _num(row[2]),
                "transactions": _num(row[3]), "closing": _num(row[4]),
            })
        if consolidated:
            data["consolidated"] = consolidated

    if SHEET_PRODUCTS in wb.sheetnames:
        for row in wb[SHEET_PRODUCTS].iter_rows(min_row=2, values_only=True):
            if not row or row[0] in (None, ""):
                continue
            key = str(row[0])
            cons = [c.strip() for c in str(row[2] or "").split("\n") if c.strip()]
            data["products"][key] = {
                "recon_sentence": str(row[1] or ""),
                "considerations": cons or list(DEFAULT_CONSIDERATIONS.get(key, [])),
                "delivery_narrative": str(row[3] or ""),
                "issued": str(row[4] or ""),
                "disp_equipment": _num(row[5]),
                "disp_other": _num(row[6]),
                "disp_transfers": _num(row[7]),
            }

    if SHEET_DELIVERIES in wb.sheetnames:
        deliveries = {k: [] for k in PRODUCT_KEYS}
        for row in wb[SHEET_DELIVERIES].iter_rows(min_row=2, values_only=True):
            if not row or row[0] in (None, ""):
                continue
            key = str(row[0])
            deliveries.setdefault(key, []).append({
                "date": str(row[1] or ""),
                "volume": _num(row[2]), "docket": _num(row[3]),
            })
        data["deliveries"] = deliveries

    return data


# --------------------------------------------------------------------------
# Calculos derivados
# --------------------------------------------------------------------------

def compute_consolidated_row(raw: dict) -> dict:
    """A partir de opening/deliveries/transactions/closing calcula el resto."""
    opening = _num(raw["opening"])
    deliveries = _num(raw["deliveries"])
    transactions = _num(raw["transactions"])
    closing = _num(raw["closing"])
    calc_stock = opening + deliveries - transactions
    net_change = closing - opening
    variance = closing - calc_stock
    pct = (variance / transactions * 100.0) if transactions else 0.0
    return {
        "site": raw["site"],
        "opening": fmt_int(opening),
        "deliveries": fmt_int(deliveries),
        "transactions": fmt_int(transactions),
        "calc_stock": fmt_int(calc_stock),
        "net_change": fmt_int(net_change),
        "closing": fmt_int(closing),
        "variance": fmt_int(variance),
        "pct": fmt_pct(pct, 2),
        "_pct_value": pct,
    }


# Constructor de "RichText" (docxtpl) con fuente Arial y negritas selectivas
# para imitar el estilo del reporte hecho a mano.
def _rt(parts) -> RichText:
    """parts: lista de (texto, bool_negrita) o (texto, bool, font_size_pt)."""
    rt = RichText()
    for item in parts:
        if len(item) == 2:
            text, bold = item
            rt.add(str(text), bold=bool(bold), font=REPORT_FONT)
        else:
            text, bold, size = item
            rt.add(str(text), bold=bool(bold), font=REPORT_FONT, size=size)
    return rt


_COUNT_WORDS = {0: "Zero", 1: "One", 2: "Two", 3: "Three", 4: "Four",
                5: "Five", 6: "Six", 7: "Seven", 8: "Eight"}


def _delivery_narrative_rt(idx: int, name: str, period: str,
                            rows: list) -> RichText:
    """Genera la narrativa de la figura de entregas como RichText, con
    inline bold para los tokens clave (cantidad, periodo, variance %)."""
    count = len(rows)
    pcts = []
    for r in rows:
        vol = _num(r["volume"])
        dock = _num(r["docket"])
        pcts.append(((dock - vol) / dock * 100.0) if dock else 0.0)
    avg = sum(pcts) / len(pcts) if pcts else 0.0
    count_word = _COUNT_WORDS.get(count, str(count))
    # Igual que en el reporte hecho a mano: singular para 0 y 1.
    deliv_word = "deliveries" if count > 1 else "delivery"
    avg_text = fmt_pct(avg, 1) if count else "0%"
    return _rt([
        (" As depicted in Figure %d. Confirmed deliveries, product " % idx,
         False),
        (name, True),
        (". ", False),
        ("%s confirmed %s" % (count_word, deliv_word), True),
        (" from ", False),
        (period, True),
        (". Average Variance across confirmed %s is " % deliv_word, False),
        (avg_text, True),
        (".", False),
    ])


def compute_delivery_rows(rows: list) -> tuple:
    """Devuelve (lista_filas_formateadas, fila_total_formateada)."""
    out, sum_vol, sum_dock = [], 0.0, 0.0
    for r in rows:
        vol = _num(r["volume"])
        dock = _num(r["docket"])
        variance = dock - vol
        pct = (variance / dock * 100.0) if dock else 0.0
        sum_vol += vol
        sum_dock += dock
        out.append({
            "date": str(r["date"]),
            "volume": fmt_int(vol), "docket": fmt_int(dock),
            "variance": fmt_int(variance), "pct": fmt_pct(pct, 1),
        })
    tot_var = sum_dock - sum_vol
    tot_pct = (tot_var / sum_dock * 100.0) if sum_dock else 0.0
    total = {
        "volume": fmt_int(sum_vol), "docket": fmt_int(sum_dock),
        "variance": fmt_int(tot_var), "pct": fmt_pct(tot_pct, 1),
    }
    return out, total


# --------------------------------------------------------------------------
# Construccion del contexto para docxtpl
# --------------------------------------------------------------------------

def build_context(data: dict, image_paths: dict, tpl: DocxTemplate) -> dict:
    """Construye el diccionario que docxtpl inyectara en la plantilla.

    data        -> estructura devuelta por default_data()/load_excel()
    image_paths -> {'fig1': 'C:/.../a.png', ...} rutas opcionales
    tpl         -> objeto DocxTemplate ya abierto (necesario para InlineImage)
    """
    period = data["period_full"]
    ctx = {"period_full": period,
           "cover_date": data.get("cover_date", "")}

    consolidated = [compute_consolidated_row(r) for r in data["consolidated"]]
    ctx["consolidated"] = consolidated

    for idx, key in enumerate(PRODUCT_KEYS, start=1):
        prefix = "p%d" % idx
        product = data["products"].get(key, {})
        name = PRODUCT_NAMES.get(key, key)

        # Fila de stock del producto = fila correspondiente de la consolidada.
        stock = next((r for r in consolidated if r["site"] == key), None)
        if stock is None:
            stock = consolidated[idx - 1] if idx - 1 < len(consolidated) else {}
        ctx["%s_stock" % prefix] = stock

        # Tabla de entregas + total.
        rows, total = compute_delivery_rows(data["deliveries"].get(key, []))
        ctx["%s_deliveries" % prefix] = rows
        ctx["%s_dtot" % prefix] = total

        # Frase de variance: RichText con Arial e inline bold (nombre del
        # producto, periodo y variance %), igual al reporte hecho a mano.
        recon_custom = (product.get("recon_sentence") or "").strip()
        if recon_custom:
            ctx["%s_recon_sentence" % prefix] = _rt([(recon_custom, False)])
        else:
            ctx["%s_recon_sentence" % prefix] = _rt([
                ("The ", False),
                (name, True),
                (" tank reconciliation for the period ", False),
                (period, True),
                (", has a total variance of ", False),
                (stock.get("pct", "0%"), True),
                (".", False),
            ])

        # Consideraciones (lista de bullets).
        cons = product.get("considerations") or list(
            DEFAULT_CONSIDERATIONS.get(key, []))
        ctx["%s_cons" % prefix] = cons

        # Narrativa de la figura de entregas (RichText con inline bold).
        narrative_custom = (product.get("delivery_narrative") or "").strip()
        if narrative_custom:
            ctx["%s_delivery_narrative" % prefix] = _rt(
                [(narrative_custom, False)])
        else:
            ctx["%s_delivery_narrative" % prefix] = _delivery_narrative_rt(
                idx, name, period, data["deliveries"].get(key, []))

        # Bloque de dispensing (RichText con inline bold).
        eq = _num(product.get("disp_equipment"))
        ot = _num(product.get("disp_other"))
        tr = _num(product.get("disp_transfers"))
        disp_total = eq + ot + tr
        issued_custom = (product.get("issued") or "").strip()
        if issued_custom:
            ctx["%s_issued" % prefix] = _rt([(issued_custom, False)])
        else:
            ctx["%s_issued" % prefix] = _rt([
                ("Issued product ", False), (name, True), (".", False)])
        ctx["%s_disp_total" % prefix] = _rt([
            ("A total ", False),
            ("Transactions", True),
            (" of ", False),
            (fmt_int(disp_total), True),
            (" L", True),
            (" was reported.", False),
        ])
        # Tres bullets de dispensing: la etiqueta y el numero van en negrita.
        ctx["%s_disp_equipment_line" % prefix] = _rt([
            ("Dispensing to Equipment: ", True),
            (fmt_int(eq), True), (" L", True), (". ", False)])
        ctx["%s_disp_other_line" % prefix] = _rt([
            ("Dispensing to other equipment ", True),
            (fmt_int(ot), True), (" L", True), (". ", False)])
        ctx["%s_disp_transfers_line" % prefix] = _rt([
            ("Transfers (Service Trucks) of ", True),
            (fmt_int(tr), True), (" L", True), (".", False)])
        # Se conservan tambien como string plano por compatibilidad.
        ctx["%s_disp_equipment" % prefix] = fmt_int(eq)
        ctx["%s_disp_other" % prefix] = fmt_int(ot)
        ctx["%s_disp_transfers" % prefix] = fmt_int(tr)

    # Figuras (imagenes). Si no hay ruta valida se inserta texto vacio.
    for fig in FIGURE_KEYS:
        path = (image_paths or {}).get(fig)
        if path and os.path.isfile(path):
            ctx[fig] = InlineImage(tpl, path, width=figure_width(fig))
        else:
            ctx[fig] = ""

    return ctx


def _build_auto_charts(data: dict, image_paths: dict, tmpdir: str) -> dict:
    """Genera con matplotlib los graficos que el software arma solo:
      - Figuras 1-4: barras de entregas (Sum of Volume vs Docket Volume).
      - Figuras 5-8: dona de distribucion de transacciones.
    Si el usuario adjunto una imagen para esa figura, se respeta la suya.
    Devuelve {fig_key: ruta_png}.
    """
    image_paths = image_paths or {}
    out = {}
    for idx, key in enumerate(PRODUCT_KEYS, start=1):
        fig_bar = "fig%d" % idx          # Figuras 1-4
        fig_pie = "fig%d" % (idx + 4)    # Figuras 5-8

        if not image_paths.get(fig_bar):
            path = os.path.join(tmpdir, "%s.png" % fig_bar)
            charts.delivery_bar_chart(data["deliveries"].get(key, []), path)
            out[fig_bar] = path

        if not image_paths.get(fig_pie):
            product = data["products"].get(key, {})
            path = os.path.join(tmpdir, "%s.png" % fig_pie)
            charts.transaction_pie_chart(
                _num(product.get("disp_equipment")),
                _num(product.get("disp_other")),
                _num(product.get("disp_transfers")), path)
            out[fig_pie] = path

    # Figuras 9-18: Tank Logs (impares) y tendencias de Reconciliation
    # (pares). Solo se autogeneran si hay datos cargados; si no, queda el
    # espacio para una imagen manual.
    tank_trends = data.get("tank_trends", {})
    safe_fill = data.get("tank_safe_fill", {})
    recon_trends = data.get("recon_trends", {})
    for i, key in enumerate(TANK_LOG_ORDER):
        fig_log = "fig%d" % (9 + 2 * i)
        fig_recon = "fig%d" % (10 + 2 * i)
        name = TANK_DISPLAY.get(key, key)

        points = tank_trends.get(key)
        if points and not image_paths.get(fig_log):
            path = os.path.join(tmpdir, "%s.png" % fig_log)
            charts.tank_log_chart(
                points, TANK_TITLES.get(key, "%s Tank Log" % name),
                path, safe_fill.get(key))
            out[fig_log] = path

        rpoints = recon_trends.get(key)
        if rpoints and not image_paths.get(fig_recon):
            path = os.path.join(tmpdir, "%s.png" % fig_recon)
            charts.recon_trend_chart(
                rpoints, "Weekly variance change %s" % name, path)
            out[fig_recon] = path
    return out


def generate_report(data: dict, image_paths: dict,
                     template_path: str, output_path: str) -> str:
    """Renderiza el reporte final. Devuelve la ruta del .docx generado.

    Las Figuras 1-8 (barras de entregas y donas de transacciones) se
    construyen automaticamente con matplotlib a partir de los datos; el
    resto de figuras (9-18 y la tabla de tareas) usan las imagenes que el
    usuario adjunte.
    """
    if not os.path.isfile(template_path):
        raise FileNotFoundError(
            "No se encontro la plantilla Word: %s" % template_path)

    tmpdir = tempfile.mkdtemp(prefix="lubes_charts_")
    try:
        auto = _build_auto_charts(data, image_paths, tmpdir)
        images = dict(image_paths or {})
        for fig_key, path in auto.items():
            images[fig_key] = path        # grafico generado por el software

        tpl = DocxTemplate(template_path)
        ctx = build_context(data, images, tpl)
        tpl.render(ctx)
        tpl.save(output_path)
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)
    return output_path
