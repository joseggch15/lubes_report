# -*- coding: utf-8 -*-
"""
Actualiza las hojas "WeeklyVariance ..." del Excel historico con la fila
de la semana actual (Date, Delivery %, Recon %) y expande el rango del
grafico de linea de cada hoja para que incluya el nuevo punto.

Productos cubiertos (los 4 con tabla consolidada en el reporte):
    S4CX30, 15W40, S5CFDM60, Tellus S3M46

Si ya existe una fila con la misma fecha, se ACTUALIZA en vez de insertar
una nueva (idempotente: el usuario puede generar el reporte varias veces
para la misma semana sin duplicar filas ni distorsionar el grafico).
"""
from __future__ import annotations

import datetime
import re

from openpyxl import load_workbook

import history
import report_model as m

# Indices 1-based de las columnas de cada hoja "WeeklyVariance ...".
WV_COL_DATE = 1
WV_COL_DELIVERY_PCT = 2
WV_COL_RECON_PCT = 3

# Solo se actualizan estos 4; S4CX10W no esta en la tabla consolidada del
# reporte, asi que no hay 'Recon %' calculado para el desde la UI.
PRODUCTS = ["S4CX30", "15W40", "S5CFDM60", "Tellus S3M46"]


def _compute_recon_pct(data: dict, key: str) -> float | None:
    """Recon % = pct de la fila consolidada (variance/transactions*100)."""
    for raw in data.get("consolidated", []):
        if raw.get("site") == key:
            computed = m.compute_consolidated_row(raw)
            return computed.get("_pct_value", 0.0)
    return None


def _compute_delivery_pct(data: dict, key: str) -> float:
    """Delivery % = % total agregado del bloque de entregas confirmadas
    de esa semana (igual al 'Grand Total Delivery Variance %' del reporte)."""
    rows = data.get("deliveries", {}).get(key, [])
    _, total = m.compute_delivery_rows(rows)
    pct_text = total.get("pct", "0%")
    # 'pct' viene formateado tipo '10,1%' -> convertir a float.
    pct_text = pct_text.replace(",", ".").rstrip("%").strip()
    try:
        return float(pct_text)
    except ValueError:
        return 0.0


def _find_row_for_date(ws, day: datetime.date) -> int | None:
    """Si ya hay una fila con esa fecha en la columna A, devuelve su numero
    (1-based). Sino, None."""
    for ri in range(3, ws.max_row + 1):
        v = ws.cell(row=ri, column=WV_COL_DATE).value
        if isinstance(v, datetime.datetime) and v.date() == day:
            return ri
        if isinstance(v, datetime.date) and v == day:
            return ri
    return None


def _write_row(ws, row: int, day: datetime.date,
               delivery_pct: float, recon_pct: float) -> None:
    ws.cell(row=row, column=WV_COL_DATE, value=datetime.datetime(
        day.year, day.month, day.day))
    ws.cell(row=row, column=WV_COL_DATE).number_format = "dd/mm/yyyy"
    ws.cell(row=row, column=WV_COL_DELIVERY_PCT, value=round(delivery_pct, 2))
    ws.cell(row=row, column=WV_COL_RECON_PCT,    value=round(recon_pct, 2))


_REF_RE = re.compile(
    r"^('?)([^'!]+)('?)!\$([A-Z]+)\$(\d+):\$([A-Z]+)\$(\d+)$")


def _extend_chart_ranges(ws, added_row: int) -> None:
    """Si la fila nueva queda fuera del rango actual de cada serie del
    grafico, extiende ese rango hasta incluirla. No mueve el inicio: la
    'ventana' del grafico simplemente crece con cada semana nueva."""
    if not ws._charts:
        return
    for chart in ws._charts:
        for ser in chart.series:
            for attr in ("cat", "val"):
                node = getattr(ser, attr, None)
                if node is None:
                    continue
                ref = (getattr(node, "numRef", None)
                       or getattr(node, "strRef", None))
                if ref is None or not ref.f:
                    continue
                mm = _REF_RE.match(ref.f)
                if not mm:
                    continue
                q1, sheet, q2, c1, r1, c2, r2 = mm.groups()
                end = int(r2)
                if added_row > end:
                    ref.f = "%s%s%s!$%s$%s:$%s$%d" % (
                        q1, sheet, q2, c1, r1, c2, added_row)


def update_for_week(excel_path: str, day: datetime.date,
                    data: dict) -> dict:
    """Para cada producto en PRODUCTS:
      1. Computa Recon % desde data['consolidated'] y Delivery % desde
         data['deliveries'].
      2. Si la hoja WeeklyVariance correspondiente ya tiene fila con la
         misma fecha -> actualiza. Sino -> agrega al final y extiende
         el rango del grafico para incluir el nuevo punto.

    Devuelve un dict resumen: {product: {'sheet', 'action', 'row',
    'delivery_pct', 'recon_pct'}}.
    """
    wb = load_workbook(excel_path)
    summary = {}
    for key in PRODUCTS:
        wv_sheet = history._match_wv_sheet(key, wb.sheetnames)
        if wv_sheet is None:
            summary[key] = {"action": "no_sheet"}
            continue
        recon_pct = _compute_recon_pct(data, key)
        if recon_pct is None:
            summary[key] = {"action": "no_consolidated_data"}
            continue
        delivery_pct = _compute_delivery_pct(data, key)
        ws = wb[wv_sheet]
        existing = _find_row_for_date(ws, day)
        if existing is not None:
            _write_row(ws, existing, day, delivery_pct, recon_pct)
            summary[key] = {"sheet": wv_sheet, "action": "updated",
                            "row": existing,
                            "delivery_pct": round(delivery_pct, 2),
                            "recon_pct": round(recon_pct, 2)}
        else:
            new_row = ws.max_row + 1
            _write_row(ws, new_row, day, delivery_pct, recon_pct)
            _extend_chart_ranges(ws, new_row)
            summary[key] = {"sheet": wv_sheet, "action": "inserted",
                            "row": new_row,
                            "delivery_pct": round(delivery_pct, 2),
                            "recon_pct": round(recon_pct, 2)}
    wb.save(excel_path)
    return summary
