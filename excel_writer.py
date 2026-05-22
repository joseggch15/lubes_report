# -*- coding: utf-8 -*-
"""
Escribe los datos extraidos de los PDFs Veridapt directamente en el Excel
historico "Reconciliation Lubs Weekly Ver2.xlsx".

Para cada producto se inserta una fila nueva (o se actualiza la existente)
en la hoja "Recon <producto>". Se respetan:

  - Valores literales: A (Date), B (Site), C (Opening), D (Deliveries),
                       H (Closing), L (To equipment), M (Other Dispenses),
                       N (Transfers).
  - Formulas calculadas: E (=SUM(L:N)), F (Calculated stock),
                         G (Net Stock change), I (Variance), J (%).
    Se traducen las referencias relativas del row anterior al nuevo row.
  - Estilos (fuente, fondo, bordes, formato numerico) clonados del row
    anterior, para que la fila nueva se vea igual al resto.
"""
from __future__ import annotations

import copy
import datetime

import openpyxl
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter

import history


# Columnas (1-indexadas para openpyxl) de la hoja "Recon <producto>".
_COL_DATE = 1          # A
_COL_SITE = 2          # B
_COL_OPENING = 3       # C
_COL_DELIVERIES = 4    # D
_COL_TRANSACTIONS = 5  # E (formula =SUM(L:N))
_COL_CALC_STOCK = 6    # F (formula)
_COL_NET_CHANGE = 7    # G (formula)
_COL_CLOSING = 8       # H
_COL_VARIANCE = 9      # I (formula)
_COL_PCT = 10          # J (formula)
_COL_TO_EQUIPMENT = 12 # L
_COL_OTHER = 13        # M
_COL_TRANSFERS = 14    # N


# Mapeo explicito (clave_PDF -> selector_hoja, tank_target).
#   selector_hoja: lo que se le pasa a history._match_sheet() para encontrar
#                  la hoja correcta. Para S4CX10W usamos 'S4CX30' porque
#                  los datos viven en la MISMA hoja 'Recon Spirax S4CX30'.
#   tank_target:   valor que ira en la columna B (Site).
#                  Nota: en el Excel del usuario aparece 'Tank1' SIN
#                  espacio para S4CX10W y 'Tank 2' / 'Tank 6' / ... CON
#                  espacio para los demas.  La fila Tank 21 es la fila
#                  resumen (formulas =Tank2+Tank1), no se toca.
PDF_TARGET = {
    "S4CX30":       {"sheet_key": "S4CX30",       "tank": "Tank 2"},
    "S4CX10W":      {"sheet_key": "S4CX30",       "tank": "Tank1"},
    "15W40":        {"sheet_key": "15W40",        "tank": "Tank 6"},
    "S5CFDM60":     {"sheet_key": "S5CFDM60",     "tank": "Tank 4"},
    "Tellus S3M46": {"sheet_key": "Tellus S3M46", "tank": "Tank 9"},
}


# Hojas que tienen una FILA RESUMEN (suma de los otros tanques).
# Cuando escribimos datos en una de estas hojas, despues de escribir los
# tanques 'fuente' se crea/actualiza automaticamente la fila resumen con
# la suma — asi el Excel queda igual a como lo manejaba antes el usuario
# (Tank 21 = Tank 2 + Tank 1 para S4CX30).
SHEET_SUMMARY_TANK = {
    "Recon Spirax S4CX30": "Tank 21",
}


def _last_data_row(ws) -> int:
    """Numero de la ultima fila con fecha valida en la columna A."""
    last = 2  # filas 1-2 son encabezados
    # Recorremos solo las primeras 16 columnas (ws.max_column suele estar
    # inflado por estilos en columnas vacias).
    for row_idx in range(3, ws.max_row + 1):
        if isinstance(ws.cell(row=row_idx, column=_COL_DATE).value,
                      datetime.datetime):
            last = row_idx
    return last


def _find_existing_row(ws, target_date: datetime.date,
                       tank_name: str | None) -> int | None:
    """Devuelve la fila si ya existe una entrada para esa fecha+tanque."""
    for row_idx in range(3, ws.max_row + 1):
        cell_date = ws.cell(row=row_idx, column=_COL_DATE).value
        if not isinstance(cell_date, datetime.datetime):
            continue
        if cell_date.date() != target_date:
            continue
        if not tank_name:
            return row_idx
        cell_site = ws.cell(row=row_idx, column=_COL_SITE).value
        if cell_site and str(cell_site).strip().lower() == tank_name.lower():
            return row_idx
    return None


def _last_row_for_tank(ws, tank_name: str | None) -> int | None:
    """Ultima fila con datos para el tanque dado (None = cualquier tanque)."""
    last = None
    for row_idx in range(3, ws.max_row + 1):
        cell_date = ws.cell(row=row_idx, column=_COL_DATE).value
        if not isinstance(cell_date, datetime.datetime):
            continue
        if tank_name is not None:
            cell_site = ws.cell(row=row_idx, column=_COL_SITE).value
            if (not cell_site or
                    str(cell_site).strip().lower() != tank_name.lower()):
                continue
        last = row_idx
    return last


def _resolve_sheet(wb, product_key: str) -> str | None:
    """Devuelve el nombre de la hoja Recon en la que vive este producto.

    Para S4CX10W devuelve 'Recon Spirax S4CX30' porque los dos tanques
    (Tank 2 = S4CX30, Tank1 = S4CX10W) viven en la misma hoja.
    """
    target = PDF_TARGET.get(product_key)
    sheet_key = target["sheet_key"] if target else product_key
    return history._match_sheet(sheet_key, wb.sheetnames)


def _resolve_tank_name(ws, product_key: str) -> str:
    """Decide el valor de la columna 'Site' (Tank) para la fila a escribir.

    Orden de prioridad:
      1. PDF_TARGET[product_key]['tank']  (mapeo explicito por producto)
      2. PRODUCT_TANK_FILTER  (fallback legacy)
      3. Tanque heredado de la fila mas reciente de la hoja
    """
    target = PDF_TARGET.get(product_key)
    if target and target.get("tank"):
        return target["tank"]
    forced = history.PRODUCT_TANK_FILTER.get(product_key)
    if forced:
        return forced
    last = _last_row_for_tank(ws, None)
    if last is not None:
        site = ws.cell(row=last, column=_COL_SITE).value
        if site:
            return str(site).strip()
    return ""


def _compute_derived(info: dict) -> dict:
    """Calcula en Python los valores que normalmente serian formulas en el
    Excel (columnas E, F, G, I, J). Asi escribimos numeros en vez de
    formulas, y openpyxl con data_only=True puede leerlos.

    Si en cambio dejaramos la formula '=SUM(L:N)' sin valor cacheado, al
    re-leer el archivo openpyxl devuelve None y la app interpreta
    Transactions = 0, perdiendo informacion. Eso es exactamente lo que
    estaba pasando con S5CFDM60 / Tellus S3M46 al recargar.
    """
    opening = float(info.get("opening", 0) or 0)
    inflow = float(info.get("inflow", 0) or 0)
    closing = float(info.get("closing", 0) or 0)
    to_eq = float(info.get("to_equipment", 0) or 0)
    other = float(info.get("other_dispenses", 0) or 0)
    trans_out = float(info.get("transfers_out", 0) or 0)

    transactions = to_eq + other + trans_out             # E = SUM(L:N)
    calc_stock = opening + inflow - transactions          # F = C+D-E
    net_change = closing - opening                        # G = H-C
    variance = closing - calc_stock                       # I = H-F
    pct = (variance / transactions) if transactions else 0.0  # J = I/E
    return {
        "transactions": transactions,
        "calc_stock": calc_stock,
        "net_change": net_change,
        "variance": variance,
        "pct": pct,
    }


def _copy_row_template(ws, source_row: int, target_row: int,
                       max_col: int = 16) -> None:
    """Clona formulas (traducidas) y estilos del row fuente al row destino.

    max_col limita el barrido a las primeras N columnas para evitar
    procesar miles de columnas vacias.
    """
    for col in range(1, max_col + 1):
        src = ws.cell(row=source_row, column=col)
        tgt = ws.cell(row=target_row, column=col)

        # Traducir formulas relativas.
        val = src.value
        if isinstance(val, str) and val.startswith("="):
            col_letter = get_column_letter(col)
            try:
                translator = Translator(
                    val, origin="%s%d" % (col_letter, source_row))
                tgt.value = translator.translate_formula(
                    "%s%d" % (col_letter, target_row))
            except Exception:
                tgt.value = val   # si falla, dejar la formula literal

        # Clonar estilos (la fila nueva se ve igual a las anteriores).
        if src.has_style:
            tgt.font = copy.copy(src.font)
            tgt.fill = copy.copy(src.fill)
            tgt.border = copy.copy(src.border)
            tgt.alignment = copy.copy(src.alignment)
            tgt.number_format = src.number_format
            tgt.protection = copy.copy(src.protection)


def write_pdf_data(excel_path: str, product_key: str,
                   period_end: datetime.date, info: dict) -> dict:
    """Inserta o actualiza una fila en la hoja Recon del producto.

    Parametros
    ----------
    excel_path  : ruta al Excel historico.
    product_key : clave del reporte (ej. '15W40').
    period_end  : fecha que se usara como la fecha de la fila.
    info        : dict con los datos parseados del PDF (de pdf_import).

    Retorna
    -------
    dict {
        'sheet':   nombre de la hoja escrita (o None si no existe),
        'row':     numero de fila escrita,
        'action':  'inserted' | 'updated' | 'no_sheet',
        'tank':    valor escrito en columna Site,
    }
    """
    wb = openpyxl.load_workbook(excel_path)
    sheet_name = _resolve_sheet(wb, product_key)
    if sheet_name is None:
        return {"sheet": None, "row": None, "action": "no_sheet", "tank": ""}

    ws = wb[sheet_name]
    tank_name = _resolve_tank_name(ws, product_key)

    existing = _find_existing_row(ws, period_end, tank_name or None)
    if existing is not None:
        target_row = existing
        action = "updated"
    else:
        # Determinar el row fuente para copiar formulas y estilos.
        source_row = (_last_row_for_tank(ws, tank_name or None)
                      or _last_row_for_tank(ws, None))
        target_row = _last_data_row(ws) + 1
        if source_row is not None:
            _copy_row_template(ws, source_row, target_row)
        action = "inserted"

    # Escribir TODOS los valores como numeros (incluyendo los que en el
    # Excel original eran formulas). Asi data_only=True puede leerlos al
    # recargar — antes las formulas devolvian None y la app mostraba
    # Transactions = 0 incorrectamente.
    derived = _compute_derived(info)
    ws.cell(row=target_row, column=_COL_DATE,
            value=datetime.datetime.combine(period_end, datetime.time()))
    ws.cell(row=target_row, column=_COL_SITE, value=tank_name)
    ws.cell(row=target_row, column=_COL_OPENING, value=info["opening"])
    ws.cell(row=target_row, column=_COL_DELIVERIES, value=info["inflow"])
    ws.cell(row=target_row, column=_COL_TRANSACTIONS,
            value=derived["transactions"])
    ws.cell(row=target_row, column=_COL_CALC_STOCK,
            value=derived["calc_stock"])
    ws.cell(row=target_row, column=_COL_NET_CHANGE,
            value=derived["net_change"])
    ws.cell(row=target_row, column=_COL_CLOSING, value=info["closing"])
    ws.cell(row=target_row, column=_COL_VARIANCE, value=derived["variance"])
    ws.cell(row=target_row, column=_COL_PCT, value=derived["pct"])
    ws.cell(row=target_row, column=_COL_TO_EQUIPMENT,
            value=info["to_equipment"])
    ws.cell(row=target_row, column=_COL_OTHER, value=info["other_dispenses"])
    ws.cell(row=target_row, column=_COL_TRANSFERS,
            value=info["transfers_out"])

    wb.save(excel_path)
    return {"sheet": sheet_name, "row": target_row, "action": action,
            "tank": tank_name}


def _upsert_summary_row(ws, day: datetime.date,
                         summary_tank: str) -> tuple | None:
    """Crea o actualiza la fila resumen (ej. Tank 21) sumando las otras
    filas del mismo dia (Tank 2 + Tank 1).

    Si no hay filas fuente para ese dia, devuelve None. Si actualiza una
    fila existente, devuelve ('updated', row). Si crea una nueva la
    inserta justo despues de la ultima fila fuente del mismo dia y
    devuelve ('inserted', row).
    """
    # 1) Recolectar todas las filas del dia.
    source_rows = []
    summary_row = None
    summary_norm = summary_tank.strip().lower()
    for r in range(3, ws.max_row + 1):
        cell_date = ws.cell(row=r, column=_COL_DATE).value
        if not isinstance(cell_date, datetime.datetime):
            continue
        if cell_date.date() != day:
            continue
        site = ws.cell(row=r, column=_COL_SITE).value
        if not site:
            continue
        if str(site).strip().lower() == summary_norm:
            summary_row = r
        else:
            source_rows.append(r)

    if not source_rows:
        return None

    # 2) Sumar columnas C, D, H, L, M, N.  Para celdas con formula sin
    # valor cacheado, asumimos 0 (no podemos calcular sin motor de
    # formulas), pero como acabamos de escribir las filas fuente con
    # valores literales, deberian estar todas computadas.
    def _sum(col):
        total = 0.0
        for r in source_rows:
            v = ws.cell(row=r, column=col).value
            if isinstance(v, (int, float)):
                total += v
        return total

    opening = _sum(_COL_OPENING)
    deliveries = _sum(_COL_DELIVERIES)
    closing = _sum(_COL_CLOSING)
    to_eq = _sum(_COL_TO_EQUIPMENT)
    other = _sum(_COL_OTHER)
    transfers = _sum(_COL_TRANSFERS)
    transactions = to_eq + other + transfers
    calc_stock = opening + deliveries - transactions
    net_change = closing - opening
    variance = closing - calc_stock
    pct = (variance / transactions) if transactions else 0.0

    # 3) Determinar si crear o actualizar.
    if summary_row is not None:
        target_row = summary_row
        action = "updated"
    else:
        # Insertar justo despues de la ultima fila fuente del dia.
        # Si esa posicion es la ultima de la hoja, no hace falta shift.
        insert_after = max(source_rows)
        last_data = _last_data_row(ws)
        if insert_after >= last_data:
            # Append al final, sin shift.
            target_row = insert_after + 1
            # Clonar estilos/formulas de una fila resumen previa si hay.
            template_row = None
            for r in range(insert_after, 2, -1):
                site = ws.cell(row=r, column=_COL_SITE).value
                if (site
                        and str(site).strip().lower() == summary_norm
                        and r != target_row):
                    template_row = r
                    break
            if template_row is None:
                template_row = insert_after  # fallback
            _copy_row_template(ws, template_row, target_row)
        else:
            # Insertar fila en el medio, shifting todo lo de abajo.
            target_row = insert_after + 1
            ws.insert_rows(target_row)
            # Buscar una fila resumen previa para clonar estilos.
            template_row = None
            for r in range(target_row - 1, 2, -1):
                site = ws.cell(row=r, column=_COL_SITE).value
                if site and str(site).strip().lower() == summary_norm:
                    template_row = r
                    break
            if template_row is not None:
                _copy_row_template(ws, template_row, target_row)
        action = "inserted"

    # 4) Escribir los valores agregados.
    ws.cell(row=target_row, column=_COL_DATE,
            value=datetime.datetime.combine(day, datetime.time()))
    ws.cell(row=target_row, column=_COL_SITE, value=summary_tank)
    ws.cell(row=target_row, column=_COL_OPENING, value=opening)
    ws.cell(row=target_row, column=_COL_DELIVERIES, value=deliveries)
    ws.cell(row=target_row, column=_COL_TRANSACTIONS, value=transactions)
    ws.cell(row=target_row, column=_COL_CALC_STOCK, value=calc_stock)
    ws.cell(row=target_row, column=_COL_NET_CHANGE, value=net_change)
    ws.cell(row=target_row, column=_COL_CLOSING, value=closing)
    ws.cell(row=target_row, column=_COL_VARIANCE, value=variance)
    ws.cell(row=target_row, column=_COL_PCT, value=pct)
    ws.cell(row=target_row, column=_COL_TO_EQUIPMENT, value=to_eq)
    ws.cell(row=target_row, column=_COL_OTHER, value=other)
    ws.cell(row=target_row, column=_COL_TRANSFERS, value=transfers)

    return action, target_row


def write_multiple(excel_path: str, items: list) -> tuple:
    """Escribe varios PDFs al Excel en una sola pasada.

    items: lista de (product_key, period_end:date, info_dict)
    Retorna (results, errors):
        results: lista de (product_key, dict resultado)
        errors:  lista de (product_key, str error)
    """
    results = []
    errors = []
    # Abrir el workbook una sola vez para todos los productos
    try:
        wb = openpyxl.load_workbook(excel_path)
    except Exception as exc:
        return [], [("(workbook)", str(exc))]

    # Hojas + dias que requieren recalcular su fila resumen al final.
    summary_targets: set = set()

    for product_key, period_end, info in items:
        try:
            sheet_name = _resolve_sheet(wb, product_key)
            if sheet_name is None:
                results.append((product_key, {
                    "sheet": None, "row": None,
                    "action": "no_sheet", "tank": ""}))
                continue

            ws = wb[sheet_name]
            tank_name = _resolve_tank_name(ws, product_key)
            existing = _find_existing_row(ws, period_end,
                                          tank_name or None)
            if existing is not None:
                target_row = existing
                action = "updated"
            else:
                source_row = (_last_row_for_tank(ws, tank_name or None)
                              or _last_row_for_tank(ws, None))
                target_row = _last_data_row(ws) + 1
                if source_row is not None:
                    _copy_row_template(ws, source_row, target_row)
                action = "inserted"

            derived = _compute_derived(info)
            ws.cell(row=target_row, column=_COL_DATE,
                    value=datetime.datetime.combine(
                        period_end, datetime.time()))
            ws.cell(row=target_row, column=_COL_SITE, value=tank_name)
            ws.cell(row=target_row, column=_COL_OPENING,
                    value=info["opening"])
            ws.cell(row=target_row, column=_COL_DELIVERIES,
                    value=info["inflow"])
            ws.cell(row=target_row, column=_COL_TRANSACTIONS,
                    value=derived["transactions"])
            ws.cell(row=target_row, column=_COL_CALC_STOCK,
                    value=derived["calc_stock"])
            ws.cell(row=target_row, column=_COL_NET_CHANGE,
                    value=derived["net_change"])
            ws.cell(row=target_row, column=_COL_CLOSING,
                    value=info["closing"])
            ws.cell(row=target_row, column=_COL_VARIANCE,
                    value=derived["variance"])
            ws.cell(row=target_row, column=_COL_PCT,
                    value=derived["pct"])
            ws.cell(row=target_row, column=_COL_TO_EQUIPMENT,
                    value=info["to_equipment"])
            ws.cell(row=target_row, column=_COL_OTHER,
                    value=info["other_dispenses"])
            ws.cell(row=target_row, column=_COL_TRANSFERS,
                    value=info["transfers_out"])

            results.append((product_key, {
                "sheet": sheet_name, "row": target_row,
                "action": action, "tank": tank_name}))

            # Si esta hoja tiene fila resumen (Tank 21 para S4CX30),
            # apuntarla para recalcular al final.
            if sheet_name in SHEET_SUMMARY_TANK:
                summary_targets.add((sheet_name, period_end))
        except Exception as exc:
            errors.append((product_key, str(exc)))

    # Crear / actualizar filas resumen (ej. Tank 21 = Tank 2 + Tank 1).
    # Se hace DESPUES de escribir todas las filas fuente para sumar los
    # ultimos valores.
    for sheet_name, day in summary_targets:
        try:
            summary_tank = SHEET_SUMMARY_TANK[sheet_name]
            ws = wb[sheet_name]
            res = _upsert_summary_row(ws, day, summary_tank)
            if res is not None:
                action, row = res
                results.append(("(summary)", {
                    "sheet": sheet_name, "row": row,
                    "action": action, "tank": summary_tank}))
        except Exception as exc:
            errors.append(("(summary %s)" % sheet_name, str(exc)))

    # Guardar todos los cambios al final.
    try:
        wb.save(excel_path)
    except Exception as exc:
        errors.append(("(save)", str(exc)))

    return results, errors
